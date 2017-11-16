#-*- coding: utf-8 -*-
from django.shortcuts import render, redirect
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from uploads.core.models import Document,CustomerData
from uploads.core.forms import DocumentForm
from openpyxl import *
from uploads.core.Process import *
from DateProcess import *
from django.utils import timezone
from django.contrib.auth.models import User


def upload(request):
    if request.method == 'POST' and request.FILES['myfile']:
        files=request.FILES.getlist('myfile')
        wb = Workbook()
        wp = CalDate()
        fs = FileSystemStorage()
        #filename = fs.save(files[0].name, files[0])
        #uploaded_file_url = fs.url(filename)
        wb2=load_workbook(files[0],data_only=True)   # 보험금청구서
        

        wb=wb2.active
        con_name=wb["B5"].value#이름
        con_number=wb["P3"].value#공제번호
        con_idennum=wb["I5"].value#주민번호
        con_date=wb["C9"].value#사고일시
        con_contract=wb["B3"].value#계약이름
        con_hospital=wb["C12"].value#병원종류
        con_type=wb["B10"].value#사고경위
        con_cop=wb["N5"].value#직장명
        con_work=wb["T5"].value#구체적으로 하는일
        con_why=wb["O9"].value#사고원인
        con_place=wb["V9"].value#사고장소
        con_dname=wb["O12"].value#진단명
        con_dcode=wb["V12"].value#질병사인코드
        con_year=wb["K28"].value#청구제출년
        con_month=wb["P28"].value#청구제출월
        con_day=wb["S28"].value#청구제출일

        car_acc=wb["I13"].value#운전여부
        car_caps=wb["V13"].value#경찰서 신고여부
        car_cop=wb["K14"].value#보험처리회사
        car_man=wb["Q14"].value#담당자
        car_tel=wb["V14"].value#연락처
        con_res=wb["W2"].value#담당자
        con_res2=wb["Y2"].value#책임자

        wb3=load_workbook(files[1],data_only=True)  #영수증
        wb=wb3.active
        rec_name=wb["B6"].value #이름
        rec_number=wb["B5"].value #등록번호
        rec_sub=wb["D5"].value #진료과
        rec_day=wb["F5"].value #진료일
        rec_kind=wb["F6"].value #진료유
        rec_pay=wb["C22"].value#본인부담총액
        rec_detail1=wb ['B8':'B19']#.value #나머지
        rec_detail2=wb['C8':'C19']
    
        print con_name


        #데이터 가공 코드
        nam=str(con_name.encode('utf-8'))
        nam2=str(rec_name.encode('utf-8'))
        con=str(con_contract.encode('utf-8'))   ##엑셀값을 가져오면 unicode로 utf-8로 바꿔줘야
        hos=str(con_hospital.encode('utf-8'))
        typ=str(con_type.encode('utf-8'))
        pay=int(rec_pay)
        
        con_date_encode=str(con_date.encode('utf-8'))

        year1=int(con_date_encode[0:4])
        month1=int(con_date_encode[7:9])
        day1=int(con_date_encode[12:14])
        year2=int(con_year)
        month2=int(con_month)
        day2=int(con_day)

        d_date=wp.deltaDate(year1,month1,day1,year2,month2,day2)

        customer=CustomerData.objects.get(name=nam)
        print customer.name



        #계약 한글->영어
        if con=="우리가족상해보장공제":
            con="Old"
        elif con=="무배당좋은이웃의료비보장공제":
            con="New"
        else:
            con="NewNew"

        if hos[:2]=="의원":
            hos="small"
        elif hos[:4]=="종합병원":
            hos="middle"
        else :
            hos="big"

        if typ=="입원":
            typ="Ipwon"
        elif typ=="통원":
            typ="Tongwon"



        # Prolog
        pr=Insurance()
        pr.tell("Contract("+con+")")
        pr.tell("Type("+typ+")")
        pr.tell("PayType(Drug)")
        value=pr.ask("Pay(a,b,c,d)")
        small=value.get(expr("a"))
        big=value.get(expr("b"))
        day=value.get(expr("c"))
        count=value.get(expr("d"))
        small=int(small)
        big=int(big)
        # #횟수 가져오기
        # if count=="Free":
        #     pass
        # else:
        #     conn = sqlite3.connect("Customer.db")
        #     cur = conn.cursor()
        #     sql="SELECT * FROM Certificate WHERE Name=?"
        #     cur.execute(sql,(nam,))
        #     if cur[2]==0:
        #         pass
        #     else:
                
            
        if pay>=small and d_date.days<=day and (customer.count>=1 or str(count)=="Free"):
            if str(count)=="Free": pass
            else:customer.count-=1;customer.save(); 
            value=customer.name+str(customer.count)+str(small)+str(big)+str(day)+str(count)
            wb4 = Workbook()
            ws2 = load_workbook("media/Report_final.xlsx")
            wb4 = ws2.active
            wb4["B3"]=con_contract#공제종목
            wb4["K3"]=con_number#공제번호
            wb4["K2"]=con_res#담당자
            wb4["M2"]=con_res2#책임자
            wb4["B5"]=con_name#성명
            wb4["E5"]=con_idennum#주민번호
            wb4["H5"]=con_cop#직장명
            wb4["K5"]=con_work#구체적으로하는일
            wb4["C6"]=con_date#사고발생일
            wb4["H6"]=con_why#사고원인
            wb4["M6"]=con_place#사고장소
            wb4["B7"]=con_type#사고경위
            wb4["C9"]=con_hospital#병원명
            wb4["H9"]=con_dname#진단명
            wb4["M9"]=con_dcode#질병사인코드
            wb4["E10"]=car_acc#운전여부
            wb4["M10"]=car_caps#경찰신고여부
            wb4["G11"]=car_cop#처리회사
            wb4["K11"]=car_man#담당자
            wb4["M11"]=car_tel#연락처

            wb4["B12"]=rec_number#등록번호
            wb4["D12"]=rec_sub#진료과
            wb4["F12"]=rec_day#진료일시
            wb4["B13"]=rec_name#이름
            wb4["F13"]=rec_kind#진료유형
            #wb4['B15':'C26']=rec_detail#나머지
            wb4["C29"]=rec_pay#본인 부담 총액
            start=15
            for row in rec_detail1:
                for cell in row:
                    if cell==None : cell.value=" ";start+=1
                    else: wb4["B"+str(start)]=cell.value;start+=1
            
            start=15
            for row in rec_detail2:
                for cell in row:
                    if cell==None : cell.value=" ";start+=1
                    else:wb4["C"+str(start)]=cell.value;start+=1


            if pay>big:
                wb4["J22"]=str(big)
                print type(big)
            elif pay<big:
                wb4["J22"]=str(pay)
            else:
                pass

            



            
            filename="Report.xlsx"
            ws2.save("media/"+filename) 
              
        #     rec_name=wb["B6"].value #이름
        # rec_number=wb["B5"].value #등록번호
        # rec_sub=wb["D5"].value #진료과
        # rec_day=wb["F5"].value #진료일
        # rec_kind=wb["F6"].value #진료유
        # rec_pay=wb["C22"].value#본인부담총액
        # rec_detail=wb ["B8:F19"].value #나머지

##            ws1.title = "Report_Sheet" 
##            ws1["A1"] = pay
##            ws1["A2"] = name
##            ws1["A3"] = con
##            filename="Report.xlsx"
##            wb4.save("media/"+filename) 
            uploaded_file_url = fs.url(filename)

            
            me = User.objects.get(username=request.user)
            obj_description=nam+"_"+con+"_"+typ+"_"+con_date_encode
            Document.objects.create(author=me, description=obj_description, document=uploaded_file_url,uploaded_at=timezone.now())
           
        else:
            return render(request, 'core/simple_upload.html')


            


        return render(request, 'core/simple_upload.html', {
            'uploaded_file_url':uploaded_file_url,
            'value':value,
        })

    return render(request, 'core/simple_upload.html')
